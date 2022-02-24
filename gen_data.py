import argparse
from openpyxl import load_workbook
from collections import defaultdict
from pygrok import Grok
from socket import gethostbyname, gaierror
from yaml import CLoader as Loader, CDumper as Dumper, load
import json

class ReadWorkbookException(Exception):
    def __init__(self, file):
        self.file = file
        self.message = "Can not read workbook: " + self.file
        super().__init__(self.message)
    def __str__(self):
        return f'{self.message}'

def read_settings_rule(rule):
    with open('./settings.yml', 'r') as f:
        try:
            return load(f.read(), Loader=Loader)['RULES'][rule]
        except:
            return False

def parse_args():
    parser = argparse.ArgumentParser()

    subparsers = parser.add_subparsers(dest='command')
    build_parser = subparsers.add_parser('build_filesystems', help='Generate CSV to build filesystems')
    patch_parser = subparsers.add_parser('map_komprise', help='Map Komprise Ids into spreadsheet')

    build_parser.add_argument("-is", "--input_spreadsheet", help="Main input Spreadsheet to use", required=True)
    build_parser.add_argument("-es", "--exports_sheet", help="Exports Workbook", required=True)
    build_parser.add_argument("-sn", "--sheet_name", help="Exports Workbook", required=True)
    build_parser.add_argument("-of", "--output_basename", help="Output files basename", required=True)
    build_parser.add_argument("-fb", "--flashblade", help="Flashblade to assign filesystems to", required=True)
    build_parser.add_argument("-ep", "--exportpolicy", help="Export Policy to assign filesystems to", required=True)
    #build_parser.add_argument("-er", "--export_rule", help="Export Policy to assign filesystems to", required=True)

    patch_parser.add_argument("-fs", "--filesystem", help="Filesystem to operate on", required=True)

    args = parser.parse_args()
    return args, parser

def tree():
    return defaultdict(tree)

class IpAddressParser:

    def __init__(self, name_address_token):
        self.name = name_address_token


    def return_type(self):

        pattern = '%{IPV4:address}'
        CIDR = {
            'CIDR': '(?<![0-9])(?:(?:[0-1]?[0-9]{1,2}|2[0-4][0-9]|25[0-5])[.](?:[0-1]?[0-9]{1,2}\
            |2[0-4][0-9]|25[0-5])[.](?:[0-1]?[0-9]{1,2}|2[0-4][0-9]|25[0-5])[.](?:[0-1]?[0-9]{1,2}\
            |2[0-4][0-9]|25[0-5]))(?![0-9])?\/[0-3]?[0-9]'
        }
        cidr_pattern = "%{CIDR:cidr}"
        grok = Grok(pattern)
        grokCidr = Grok(cidr_pattern, custom_patterns=CIDR)

        netaddr = None
        addr = None
        if self.name:
            ret = grokCidr.match(self.name)
        else:
            self.name = ''
            ret = grokCidr.match(self.name)

        try:
            netaddr = ret['cidr']
        except:
            netaddr = None
            ret = grok.match(self.name)
            try:
                addr = ret['address']
            except:
                addr = None

        if not addr and not netaddr:
            return 'NAME'
        if addr:
            return 'ADDRESS'
        if netaddr:
            return 'CIDR'

    def return_entry(self):

        if self.return_type() == 'NAME':
            try:
                ret = ip_address = gethostbyname(self.name)
            except gaierror:
                ret = self.name
            return ret

        if self.return_type() == 'CIDR':
            #net = IPNetwork(self.name)
            #hosts = [ str(host) for host in net.iter_hosts()]
            #print(hosts)
            return self.name

        if self.return_type() == 'ADDRESS':
            return self.name


class Rule:

    def __init__(self, ro, rw, superuser):
        self.rw = 'ro'
        self.root_squash = 'root_squash'
        self.fileid_32bit = 'no_fileid_32bit'

        if rw == 'any' and ro == 'any':
            self.rw = 'rw'
        if rw == 'none' and ro == 'any':
            self.rw = 'ro'
        if superuser == 'any':
            self.root_squash = 'no_root_squash'

    def ret_settings(self):

        if self.fileid_32bit == 'fileid-32bit' :
            return self.rw+ ',' +self.root_squash+ ',' +self.fileid_32bit
        else:
            return self.rw+ ',' +self.root_squash


class ExportWorkBookHandler:

    def __init__(self, file, sheet):
        self.wb = load_workbook(file)
        self.ws = self.wb[sheet]

    def locate_policy(self, policy):

        self.fs_to_export_rule = {}
        found = False
        trules = dict(self.final_rules)
        for vserver in trules:
            try:
                if len(trules[vserver][policy].keys()) > 0:
                    found = True
                    return trules[vserver][policy]
            except:
                pass
        if not found:
            return None

    def parse_sheet(self):
        policy_rules = tree()
        for row in self.ws:

            try:
                policy_rules[row[0].value][row[1].value].append({
                    #'client': row[4].value,
                    'client': IpAddressParser(row[4].value).return_entry(),
                    'rorule': row[5].value,
                    'rwrule': row[6].value,
                    'anon': row[7].value,
                    'superuser': row[8].value,
                    'allow_superuser': row[9].value,
                    'allow_dev': row[10].value
                })
            except:
                policy_rules[row[0].value][row[1].value] = []
                policy_rules[row[0].value][row[1].value].append({
                    #'client': row[4].value,
                    'client': IpAddressParser(row[4].value).return_entry(),
                    'rorule': row[5].value,
                    'rwrule': row[6].value,
                    'anon': row[7].value,
                    'superuser': row[8].value,
                    'allow_superuser': row[9].value,
                    'allow_dev': row[10].value
                })
        self.rules = policy_rules

    def xlate_rules(self):
        for vserver in self.rules:
            for policy in self.rules[vserver]:
                for rule in self.rules[vserver][policy]:
                    tRule = Rule(rule['rorule'], rule['rwrule'], rule['superuser'])
                    rule.update({
                        'rule': tRule.ret_settings()
                    })

        self.final_rules = tree()
        for vserver in self.rules:
            for policy in self.rules[vserver]:
                self.final_rules[vserver][policy] = {}
                for rule in self.rules[vserver][policy]:
                    try:
                        self.final_rules[vserver][policy][rule['rule']].append(rule['client'])
                    except:
                        self.final_rules[vserver][policy][rule['rule']] = []
                        self.final_rules[vserver][policy][rule['rule']].append(rule['client'])

    def print_rules(self):
        print(json.dumps(self.final_rules, indent=4))


class WorkBookHandler:

    def __init__(self, wb, saved_name, fb, ep, build_file):
        try:
            self.saved_name = saved_name
            self.wb = load_workbook(wb)
            self.ws = self.wb.active
            self.fb = fb
            self.ep = ep
            self.build_file = build_file
        except:
            raise ReadWorkbookException(wb)

    def save_wb(self):
        self.wb.save(self.saved_name)

    def gen_export_tool_input(self, output_base, rule):
        komprise_rules = read_settings_rule(rule)
        output_file = output_base + '_exports.json'
        f = open(output_file, 'w')
        self.t_output = {}
        self.t_output[self.fb] = []
        for fs in self.fs_to_policy:
            trule = {}
            if fs:
                if isinstance(self.fs_to_policy[fs], dict):
                    trule['name'] = fs
                    trule['rules'] = komprise_rules + ' '
                    for rule in self.fs_to_policy[fs]:
                        trule['rules'] = trule['rules'] + '-' + rule + ' ' + ' '.join(self.fs_to_policy[fs][rule]) + ' '
            self.t_output[self.fb].append(trule)
        f.write(json.dumps(self.t_output, indent=4))
        f.close()


    def match_policy(self, policy_class):
        self.fs_to_policy = {}
        for row in self.ws:
            #print("POLICY: " + row[6].value)
            self.fs_to_policy[row[24].value] = policy_class.locate_policy(row[6].value)

        print(json.dumps(self.fs_to_policy, indent=4))

    def gen_input_rules(self):
        self.input_rules = {}
        self.input_rules[self.fb] = []

        for each in self.fs_to_policy:
            self.input_rules[self.fb].append({
                'name': each,
                'rules': ''
            })

        return json.dumps(self.input_rules, indent=4)


    def gen_names_and_sizes(self):

        f = open(self.build_file, 'w')

        count = 1
        for row in self.ws:
            col_s = row[18].value
            col_e = row[4].value
            col_r = row[17].value
            col_p = row[15].value
            try:
                bu = row[22].value.replace(' ', '').lower()
            except:
                row[22].value = ''

            qtree = row[14].value
            volume = row[2].value

            if qtree == '-' or qtree == '':
                # You're a volume
                if 'GB' in col_s:
                    col_s = col_s.replace('GB', '')
                    col_s = float(col_s)
                if 'MB' in str(col_s):
                    col_s = col_s.replace('MB', '')
                    col_s = float(col_s)/1024

                try:
                    if col_s > 0:
                        size = int(col_e.replace('GB', ''))
                        if 'MB' in col_e:
                            size = float(col_e.replace('MB', '')) / 1024
                    else:
                        size = int(col_r.replace('GB', ''))
                        if 'MB' in col_r:
                            size = float(col_r.replace('MB', '')) / 1024
                except:
                    size = 0
                if bu != '' and bu != '-':
                    volume = bu + '_' + volume
                f.write(self.fb+ ',' +str(size)+ ',' +self.ep+ ',' +volume+ ',,,\n')
                #print(self.fb+ ',' +str(size)+ ',' +self.ep+ ',' +volume+ ',,,')
                self.ws.cell(row=count, column=25, value=volume)
                self.ws.cell(row=count, column=26, value=size)

            else:
                # You're a qtree
                vname = volume + '_' + qtree

                try:
                    if 'MB' in col_p:
                        size = float(col_p.replace('MB','')) / 1024
                    elif 'GB' in col_p:
                        size = float(col_p.replace('GB', ''))
                    else:
                        size = float(col_p)
                except:
                    size = 0.0

                if bu != '' and bu != '-':
                    volume = bu + '_' + volume


                f.write(self.fb+ ',' +str(size)+ ',' +self.ep+ ',' +volume+ ',,,\n')
                #print(self.fb+ ',' +str(size)+ ',' +self.ep+ ',' +volume+ ',,,')
                self.ws.cell(row=count, column=25, value=volume)
                self.ws.cell(row=count, column=26, value=size)
            count += 1

class RequestHandler:

    def __init__(self, args, parser):
        self.args = args
        self.parser = parser

    def exec_request(self):

        if self.args.command == 'build_filesystems':

            new_name = self.args.output_basename + '_saved_sheet.xlsx'
            build_file = self.args.output_basename + '_fs_build.csv'

            self.wb = WorkBookHandler(self.args.input_spreadsheet, new_name, self.args.flashblade, self.args.exportpolicy, build_file)
            self.wb.gen_names_and_sizes()
            self.wb.save_wb()

    def match_policies(self, myPolicy):
        self.wb.match_policy(myPolicy)

    def ret_rules(self):
        print(self.wb.gen_input_rules())

    def write_input(self):
        file = self.args.output_basename
        self.wb.gen_export_tool_input(file, self.args.exportpolicy)


def main():
    args, parser = parse_args()
    requestHandler = RequestHandler(args, parser)
    requestHandler.exec_request()

    exportWorkBook = ExportWorkBookHandler(args.exports_sheet, args.sheet_name)
    exportWorkBook.parse_sheet()
    exportWorkBook.xlate_rules()
    exportWorkBook.print_rules()
    requestHandler.match_policies(exportWorkBook)
    requestHandler.write_input()


if __name__ == "__main__":
    main()